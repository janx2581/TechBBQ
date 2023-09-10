# Import necessary libraries
import streamlit as st
import replicate
import os

import pandas as pd
from openpyxl import load_workbook, Workbook # Import openpyxl

# Initialize an empty DataFrame
data = []

# Set the Replicate API token
os.environ["REPLICATE_API_TOKEN"] = "r8_bBCxqkUHTfwz818CIEtNtnojCT7yRZJ3gytkY"

# Set the title of the app
st.set_page_config(page_title="HTHC AI 🔴⚪️")

# Add custom CSS styling to the app
st.markdown("""
    <style>
        .stApp { background-color: #000000; }
        body { font-family: 'Times New Roman', sans-serif; line-height: 1.5; color: #FFFFFF; }
        .stTextInput input { background-color: #000000; color: #FFFFFF; border: 1px solid #FF0808; }
        .stButton>button { background-color: #FF0808; border: none; }
        .st-chat-messages { border: none; }
    </style>
    """, unsafe_allow_html=True)

# Create a sidebar with some information
with st.sidebar:
    st.title('Welcome to the HTHC AI chatbot 🔴⚪️')
    st.markdown('It is called Health Tech Hygge AI')
    st.markdown('📖 Learn more about Health Tech Hub Copenhagen [here](https://healthtechhub.org/)!')

# Store LLM generated responses in session state
if "messages" not in st.session_state.keys():
    st.session_state.messages = [{"role": "content creator", "content": "Let me help you be creative"}]

# Display or clear chat messages
for message in st.session_state.messages:
    with st.chat_message(message["role"]):
        st.write(message["content"])


# Function to clear chat history
def clear_chat_history():
    st.session_state.messages = [{"role": "assistant", "content": "Let me help you be creative"}]

# Add a button to clear chat history in the sidebar
st.sidebar.button('Clear Chat History', on_click=clear_chat_history)

# Function for generating LLaMA2 response
def generate_llama2_response(prompt_input):
    # Build the string dialogue by combining user and assistant messages
    string_dialogue = "You are a creative content creator from Health Tech Hub Copenhagen. You do not respond as 'User' or pretend to be 'User'. You only respond once as 'Assistant'. Health Tech Hub Copenhagen's catchphrase is Making Health Tech Everyones Business"
    for dict_message in st.session_state.messages:
        if dict_message["role"] == "user":
            string_dialogue += "User: " + dict_message["content"] + "\n\n"
        else:
            string_dialogue += "Assistant: " + dict_message["content"] + "\n\n"

    # Generate LLaMA2 response using the replicate.run() function
    output = replicate.run(
        # 'a16z-infra/llama13b-v2-chat:df7690f1994d94e96ad9d568eac121aecf50684a0b0963b25a41cc40061269e5', # The 13B parameter model
        "meta/llama-2-70b-chat:35042c9a33ac8fd5e29e27fb3197f33aa483f72c2ce3b0b9d201155c7fd2a287", # The 70B parameter model
        input={"prompt": f"{string_dialogue} {prompt_input} Assistant: ",
               "temperature": 0.1, "top_p": 0.9, "max_length": 512, "repetition_penalty": 1})
    return output

# Get user input prompt using st.chat_input()
if prompt := st.chat_input(disabled=not os.environ['REPLICATE_API_TOKEN']):
    # Add user message to session state
    st.session_state.messages.append({"role": "user", "content": prompt})
    with st.chat_message("user"):
        st.write(prompt)

# Generate a new response if the last message is not from the assistant
if st.session_state.messages[-1]["role"] != "assistant":
    with st.chat_message("Assistant"):
        with st.spinner("Thinking..."):
            # Generate LLaMA2 response using the generate_llama2_response() function
            response = generate_llama2_response(prompt)
            placeholder = st.empty()
            full_response = ''
            for item in response:
                full_response += item
                placeholder.markdown(full_response)
            placeholder.markdown(full_response)
    message = {"role": "assistant", "content": full_response}
    st.session_state.messages.append(message)
    st.session_state['latest_response'] = full_response  # Store the latest response in session state

# Check if the form should be visible
if 'show_form' not in st.session_state:
    st.session_state['show_form'] = False

# Create a button to toggle the visibility of the form
if st.button('Ready to submit? Click here'):
    st.session_state['show_form'] = not st.session_state['show_form']

# Placeholder to hold the form
form_placeholder = st.empty()

# If the button has been pressed, show the form
if st.session_state['show_form']:
    with form_placeholder.form(key='entry_form'):
        st.header("please fill out below")
        participant_name = st.text_input("Name")
        participant_email = st.text_input("Email")
        participant_entry = st.text_area("Your Creative Text", value=st.session_state.get('latest_response', ''))
        submit_button = st.form_submit_button(label='Submit Entry')

        if submit_button:
            data_dict = {'Name': participant_name, 'Email': participant_email, 'Creative Text': participant_entry}

            # Check if the Excel file already exists
            if os.path.exists('entries.xlsx'):
                # Load the existing workbook and select the active worksheet
                book = load_workbook('entries.xlsx')
                sheet = book.active

                # Find the next empty row to write the data
                next_row = sheet.max_row + 1
            else:
                # Create a new workbook and worksheet
                book = Workbook()
                sheet = book.active

                # Write the headers in the first row
                sheet['A1'] = 'Name'
                sheet['B1'] = 'Email'
                sheet['C1'] = 'Creative Text'
                next_row = 2

            # Write the new entry in the next empty row
            sheet[f'A{next_row}'] = data_dict['Name']
            sheet[f'B{next_row}'] = data_dict['Email']
            sheet[f'C{next_row}'] = data_dict['Creative Text']

            # Save the workbook
            book.save('entries.xlsx')
            st.success('Entry submitted and saved to entries.xlsx')
